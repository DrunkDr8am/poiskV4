import os
import fnmatch
import zipfile
import tempfile
from io import BytesIO
from typing import Set, Dict, List
import logging

# Глобальная переменная для хранения ключевых слов в нижнем регистре
KEYWORDS_LOWER = set()


def load_keywords(keywords_file: str) -> List[str]:
    """Загрузка ключевых слов из файла с проверкой кодировки"""
    global KEYWORDS_LOWER
    encodings = ['utf-8', 'cp1251', 'iso-8859-1', 'utf-8-sig']
    for encoding in encodings:
        try:
            with open(keywords_file, 'r', encoding=encoding) as f:
                keywords = [line.strip() for line in f if line.strip()]
                if keywords:
                    # Сохраняем ключевые слова в нижнем регистре для быстрого поиска
                    KEYWORDS_LOWER = {kw.lower() for kw in keywords}
                    return keywords
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Не удалось декодировать файл {keywords_file} с поддержанными кодировками: {encodings}")


def search_in_text(text: str) -> Set[str]:
    """Поиск ключевых слов в тексте с использованием множеств для скорости"""
    if not text:
        return set()

    text_lower = text.lower()
    return {kw for kw in KEYWORDS_LOWER if kw in text_lower}


def search_in_image(image_data: BytesIO or str, config: dict) -> Set[str]:
    """Распознавание текста с изображения"""
    # Проверяем доступность OCR через конфиг
    if not config.get('has_ocr', False):
        return set()

    try:
        from PIL import Image
        import pytesseract
    except ImportError:
        return set()

    try:
        img = Image.open(image_data) if isinstance(image_data, BytesIO) else Image.open(image_data)

        if img.mode not in ('RGB', 'L'):
            img = img.convert('RGB')

        # Используем настройки из конфига
        languages = config.get('tesseract_languages', 'rus')
        config_param = config.get('tesseract_config', '--oem 3 --psm 6')

        text = pytesseract.image_to_string(img, lang=languages, config=config_param)
        return search_in_text(text)
    except Exception as e:
        logging.error(f"Ошибка обработки изображения: {e}")
        return set()


def search_in_pdf(pdf_path: str, config: dict) -> Set[str]:
    """Обработка PDF файлов"""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return set()

    found = set()
    try:
        with fitz.open(pdf_path) as doc:
            for page in doc:
                # Текст со страницы
                text = page.get_text()
                found.update(search_in_text(text))

                # Обработка изображений (только если есть OCR)
                for img in page.get_images(full=True):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    if base_image and "image" in base_image:
                        image_data = BytesIO(base_image["image"])
                        found.update(search_in_image(image_data, config))
    except Exception as e:
        logging.error(f"Ошибка обработки PDF {pdf_path}: {e}")
    return found


def search_in_docx(docx_path: str, config: dict) -> Set[str]:
    """Обработка DOCX файлов"""
    try:
        import docx2txt
    except ImportError:
        return set()

    found = set()
    try:
        # Текст из документа
        text = docx2txt.process(docx_path)
        found.update(search_in_text(text))

        # Изображения из документа (только если есть OCR)
        with tempfile.TemporaryDirectory() as temp_dir:
            docx2txt.process(docx_path, temp_dir)
            for img_file in os.listdir(temp_dir):
                if img_file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')):
                    img_path = os.path.join(temp_dir, img_file)
                    found.update(search_in_image(img_path, config))
    except Exception as e:
        logging.error(f"Ошибка обработки DOCX {docx_path}: {e}")
    return found


def search_in_excel(excel_path: str) -> Set[str]:
    """Обработка Excel файлов"""
    try:
        import pandas as pd
        import openpyxl
    except ImportError:
        return set()

    found = set()
    try:
        # Пропускаем временные файлы Excel
        if os.path.basename(excel_path).startswith('~$'):
            return set()

        if excel_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            found.update(search_in_text(cell))
        else:  # .xls
            df = pd.read_excel(excel_path, sheet_name=None)
            for sheet_name, sheet_data in df.items():
                for _, row in sheet_data.iterrows():
                    for value in row:
                        if isinstance(value, str):
                            found.update(search_in_text(str(value)))
    except Exception as e:
        logging.error(f"Ошибка обработки Excel {excel_path}: {e}")
    return found


def search_in_archive(archive_path: str, extensions: List[str], config: dict) -> Set[str]:
    """Обработка архивов с поддержкой изображений"""
    found = set()
    try:
        if archive_path.endswith('.zip'):
            with zipfile.ZipFile(archive_path, 'r') as z:
                for file in z.namelist():
                    if any(fnmatch.fnmatch(file, ext) for ext in extensions):
                        # Для текстовых файлов читаем напрямую
                        if file.lower().endswith(('.txt', '.csv', '.log', '.xml', '.html', '.htm')):
                            with z.open(file) as f:
                                content = f.read().decode('utf-8', errors='ignore')
                                found.update(search_in_text(content))
                        # Для изображений и других бинарных файлов извлекаем во временную директорию
                        else:
                            with tempfile.TemporaryDirectory() as temp_dir:
                                z.extract(file, temp_dir)
                                extracted_file = os.path.join(temp_dir, file)
                                if os.path.isfile(extracted_file):
                                    # Обрабатываем изображения
                                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')):
                                        found.update(search_in_image(extracted_file, config))
                                    # Обрабатываем PDF
                                    elif file.lower().endswith('.pdf') and config.get('has_pdf', False):
                                        found.update(search_in_pdf(extracted_file, config))
                                    # Обрабатываем DOCX
                                    elif file.lower().endswith('.docx') and config.get('has_docx', False):
                                        found.update(search_in_docx(extracted_file, config))
                                    # Обрабатываем Excel
                                    elif file.lower().endswith(('.xls', '.xlsx')) and config.get('has_excel', False):
                                        found.update(search_in_excel(extracted_file))


        elif archive_path.endswith('.7z'):
            try:
                import py7zr
            except ImportError:
                return set()
            with tempfile.TemporaryDirectory() as temp_dir:
                try:
                    with py7zr.SevenZipFile(archive_path, mode='r') as z:
                        # Извлекаем все файлы
                        z.extractall(path=temp_dir)
                    # Рекурсивно обходим извлеченные файлы
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, temp_dir)
                            if any(fnmatch.fnmatch(relative_path, ext) for ext in extensions):
                                # Обрабатываем файлы в зависимости от типа
                                if file.lower().endswith(('.txt', '.csv', '.log', '.xml', '.html', '.htm')):
                                    try:
                                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                            content = f.read()
                                            found.update(search_in_text(content))
                                    except:
                                        try:
                                            with open(file_path, 'rb') as f:
                                                content = f.read().decode('utf-8', errors='ignore')
                                                found.update(search_in_text(content))
                                        except:
                                            pass
                                elif file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')):
                                    found.update(search_in_image(file_path, config))
                                elif file.lower().endswith('.pdf') and config.get('has_pdf', False):
                                    found.update(search_in_pdf(file_path, config))
                                elif file.lower().endswith('.docx') and config.get('has_docx', False):
                                    found.update(search_in_docx(file_path, config))
                                elif file.lower().endswith(('.xls', '.xlsx')) and config.get('has_excel', False):
                                    found.update(search_in_excel(file_path))
                except Exception as e:
                    logging.error(f"Ошибка обработки 7z архива {archive_path}: {e}")

        elif archive_path.endswith('.rar'):
            try:
                import rarfile
            except ImportError:
                return set()

            with rarfile.RarFile(archive_path, 'r') as z:
                for file in z.namelist():
                    if any(fnmatch.fnmatch(file, ext) for ext in extensions):
                        # Для текстовых файлов читаем напрямую
                        if file.lower().endswith(('.txt', '.csv', '.log', '.xml', '.html', '.htm')):
                            with z.open(file) as f:
                                content = f.read().decode('utf-8', errors='ignore')
                                found.update(search_in_text(content))
                        # Для изображений и других бинарных файлов извлекаем во временную директорию
                        else:
                            with tempfile.TemporaryDirectory() as temp_dir:
                                z.extract(file, temp_dir)
                                extracted_file = os.path.join(temp_dir, file)
                                if os.path.isfile(extracted_file):
                                    # Обрабатываем изображения
                                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')):
                                        found.update(search_in_image(extracted_file, config))
                                    # Обрабатываем PDF
                                    elif file.lower().endswith('.pdf') and config.get('has_pdf', False):
                                        found.update(search_in_pdf(extracted_file, config))
                                    # Обрабатываем DOCX
                                    elif file.lower().endswith('.docx') and config.get('has_docx', False):
                                        found.update(search_in_docx(extracted_file, config))
                                    # Обрабатываем Excel
                                    elif file.lower().endswith(('.xls', '.xlsx')) and config.get('has_excel', False):
                                        found.update(search_in_excel(extracted_file))

    except Exception as e:
        logging.error(f"Ошибка обработки архива {archive_path}: {e}")
    return found


def process_file(file_path: str, extensions: List[str], max_file_size: int, config: dict) -> Dict[str, Set[str]]:
    """Обработка отдельного файла"""
    found = set()
    try:
        # Проверяем размер файла
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        if file_size_mb > max_file_size:
            logging.warning(
                f"Пропуск файла {file_path} (размер {file_size_mb:.2f} МБ превышает лимит {max_file_size} МБ)")
            return {}

        ext = os.path.splitext(file_path)[1].lower()

        # Проверяем, соответствует ли файл заданным расширениям
        if not any(fnmatch.fnmatch(file_path, ext_pattern) for ext_pattern in extensions):
            return {}

        # Обработка в зависимости от типа файла
        if ext in ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff'):
            # Проверяем доступность OCR через конфиг
            if config.get('has_ocr', False):
                found = search_in_image(file_path, config)
            else:
                logging.info(f"Пропуск изображения {file_path} (OCR недоступен)")
        elif ext == '.pdf':
            # Проверяем доступность обработки PDF
            if config.get('has_pdf', False):
                found = search_in_pdf(file_path, config)
            else:
                logging.info(f"Пропуск PDF {file_path} (обработка PDF недоступна)")
        elif ext == '.docx':
            # Проверяем доступность обработки DOCX
            if config.get('has_docx', False):
                found = search_in_docx(file_path, config)
            else:
                logging.info(f"Пропуск DOCX {file_path} (обработка DOCX недоступна)")
        elif ext in ('.xls', '.xlsx'):
            # Проверяем доступность обработки Excel
            if config.get('has_excel', False):
                found = search_in_excel(file_path)
            else:
                logging.info(f"Пропуск Excel {file_path} (обработка Excel недоступна)")
        elif ext in ('.zip', '.7z', '.rar'):
            # Для архивов проверяем доступность соответствующих модулей
            if ext == '.7z' and not config.get('has_7z', False):
                logging.info(f"Пропуск 7Z {file_path} (обработка 7Z недоступна)")
            elif ext == '.rar' and not config.get('has_rar', False):
                logging.info(f"Пропуск RAR {file_path} (обработка RAR недоступна)")
            else:
                found = search_in_archive(file_path, extensions, config)  # Передаем config
        else:
            # Обработка текстовых файлов
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    found = search_in_text(content)
            except UnicodeDecodeError:
                # Если UTF-8 не работает, пробуем другие кодировки
                encodings = ['cp1251', 'iso-8859-1', 'latin1']
                for encoding in encodings:
                    try:
                        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                            content = f.read()
                            found = search_in_text(content)
                            break
                    except UnicodeDecodeError:
                        continue

        return {file_path: found} if found else {}
    except Exception as e:
        logging.error(f"Ошибка обработки файла {file_path}: {e}")
        return {}
