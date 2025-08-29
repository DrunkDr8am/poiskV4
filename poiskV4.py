import os
import fnmatch
import sys
import tempfile
import zipfile
import requests
from pathlib import Path
import warnings
from io import BytesIO
from typing import List, Dict, Set
from concurrent.futures import ThreadPoolExecutor, as_completed
import configparser
from tqdm import tqdm
import time
import logging
import shutil


def setup_portable_tesseract():
    """Настройка портативного Tesseract OCR"""
    try:
        # Определяем базовый путь в зависимости от того, запущено ли как exe
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))

        tesseract_path = os.path.join(base_path, "tesseract", "tesseract.exe")
        tessdata_path = os.path.join(base_path, "tesseract", "tessdata")

        # Проверяем, существует ли портативный Tesseract
        if os.path.exists(tesseract_path) and os.path.exists(tessdata_path):
            # Устанавливаем путь к Tesseract
            pytesseract.pytesseract.tesseract_cmd = tesseract_path

            # Устанавливаем путь к данным
            os.environ['TESSDATA_PREFIX'] = tessdata_path

            # Проверяем, работает ли Tesseract
            try:
                version = pytesseract.get_tesseract_version()
                logging.info(f"Портативный Tesseract OCR найден: версия {version}")
                return True
            except Exception as e:
                logging.error(f"Портативный Tesseract найден, но не работает корректно: {e}")
                return False
        else:
            logging.warning("Портативный Tesseract не найден или неполная установка")
            return False
    except Exception as e:
        logging.error(f"Ошибка при настройке портативного Tesseract: {e}")
        return False


# Настройка логирования с очисткой предыдущих логов
def setup_logging(log_file='search_log.txt'):
    # Удаляем предыдущий файл логов, если он существует
    if os.path.exists(log_file):
        try:
            os.remove(log_file)
        except Exception as e:
            print(f"Не удалось удалить старый файл логов: {e}")

    # Создаем новый логгер с очисткой предыдущих обработчиков
    logger = logging.getLogger()
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    # Настраиваем логирование
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, mode='w', encoding='utf-8'),  # 'w' для перезаписи файла
            logging.StreamHandler(sys.stdout)
        ]
    )


# Попытаемся импортировать опциональные зависимости
try:
    import py7zr

    HAS_7Z = True
except ImportError:
    HAS_7Z = False
    print("Модуль py7zr не установлен. Поддержка 7z архивов отключена.")

try:
    import rarfile

    HAS_RAR = True
except ImportError:
    HAS_RAR = False
    print("Модуль rarfile не установлен. Поддержка RAR архивов отключена.")

try:
    import fitz  # PyMuPDF

    HAS_PDF = True
except ImportError:
    HAS_PDF = False
    print("Модуль PyMuPDF не установлен. Поддержка PDF отключена.")

# Проверяем и устанавливаем Tesseract при необходимости
# Проверяем доступность OCR
HAS_OCR = False
try:
    from PIL import Image
    import pytesseract

    # Используем только портативный Tesseract
    HAS_OCR = setup_portable_tesseract()

    if not HAS_OCR:
        print("Портативный Tesseract не найден или не работает. Поиск по изображениям отключен.")

except ImportError:
    print("Модули для OCR не установлены. Поиск по изображениям отключен.")
    HAS_OCR = False

try:
    import docx2txt

    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    print("Модули для DOCX не установлены. Поддержка DOCX отключена.")

try:
    import pandas as pd
    import openpyxl

    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("Модули для Excel не установлены. Поддержка Excel отключена.")

warnings.filterwarnings("ignore")

# Глобальная переменная для хранения ключевых слов в нижнем регистре
KEYWORDS_LOWER = set()


def load_config(config_file="config.txt"):
    """Загрузка конфигурации из файла"""
    config = configparser.ConfigParser()

    # Значения по умолчанию
    defaults = {
        'extensions': ['*.txt', '*.pdf', '*.docx', '*.xlsx', '*.jpg', '*.png', '*.zip', '*.rar', '*.7z'],
        'keywords_file': 'keywords.txt',
        'directory': '.',
        'threads': '4',
        'output_file': 'search_results.txt',
        'search_images': 'false',
        'max_file_size': '50',
        'log_file': 'search_log.txt',
        'tesseract_languages': 'rus',
        'tesseract_config': '--oem 3 --psm 6'
    }

    # Если файл конфигурации существует, загружаем его
    if os.path.exists(config_file):
        try:
            config.read(config_file, encoding='utf-8')
        except Exception as e:
            print(f"Ошибка чтения конфигурационного файла: {e}")
            return defaults

    # Получаем значения из конфига или используем значения по умолчанию
    extensions = config.get('Settings', 'extensions', fallback=','.join(defaults['extensions'])).split(',')
    keywords_file = config.get('Settings', 'keywords_file', fallback=defaults['keywords_file'])
    directory = config.get('Settings', 'directory', fallback=defaults['directory'])
    threads = config.getint('Settings', 'threads', fallback=int(defaults['threads']))
    output_file = config.get('Settings', 'output_file', fallback=defaults['output_file'])
    search_images = config.getboolean('Settings', 'search_images', fallback=False)
    max_file_size = config.getint('Settings', 'max_file_size', fallback=int(defaults['max_file_size']))
    log_file = config.get('Settings', 'log_file', fallback=defaults['log_file'])
    tesseract_languages = config.get('Settings', 'tesseract_languages', fallback=defaults['tesseract_languages'])
    tesseract_config = config.get('Settings', 'tesseract_config', fallback=defaults['tesseract_config'])

    # Очищаем значения от пробелов
    extensions = [ext.strip() for ext in extensions]
    keywords_file = keywords_file.strip()
    directory = directory.strip()
    output_file = output_file.strip()
    log_file = log_file.strip()

    # Если поиск по изображениям отключен, убираем изображения из расширений
    if not search_images:
        extensions = [ext for ext in extensions if
                      not ext.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff'))]

    return {
        'extensions': extensions,
        'keywords_file': keywords_file,
        'directory': directory,
        'threads': threads,
        'output_file': output_file,
        'search_images': search_images,
        'max_file_size': max_file_size,
        'log_file': log_file,
        'tesseract_languages': tesseract_languages,
        'tesseract_config': tesseract_config
    }


def load_keywords(keywords_file: str) -> List[str]:
    """Загрузка ключевых слов из файла с проверкой кодировки"""
    global KEYWORDS_LOWER
    encodings = ['utf-8', 'cp1251', 'iso-8859-1', 'utf-8-sig']
    for encoding in encodings:
        try:
            with open(keywords_file, 'r', encoding=encoding) as f:
                keywords = [line.strip() for line in f if line.strip()]
                if keywords:
                    # Сохраняем ключевые слова в нижнем регистре для быстрого поиска
                    KEYWORDS_LOWER = {kw.lower() for kw in keywords}
                    return keywords
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Не удалось декодировать файл {keywords_file} с поддержанными кодировками: {encodings}")


def search_in_text(text: str) -> Set[str]:
    """Поиск ключевых слов в тексте с использованием множеств для скорости"""
    if not text:
        return set()

    text_lower = text.lower()
    return {kw for kw in KEYWORDS_LOWER if kw in text_lower}


def search_in_image(image_data: BytesIO or str, config: dict) -> Set[str]:
    """Распознавание текста с изображения"""
    if not HAS_OCR:
        return set()

    try:
        img = Image.open(image_data) if isinstance(image_data, BytesIO) else Image.open(image_data)

        if img.mode not in ('RGB', 'L'):
            img = img.convert('RGB')

        # Используем настройки из конфига
        languages = config.get('tesseract_languages', 'rus')
        config_param = config.get('tesseract_config', '--oem 3 --psm 6')

        text = pytesseract.image_to_string(img, lang=languages, config=config_param)
        return search_in_text(text)
    except Exception as e:
        logging.error(f"Ошибка обработки изображения: {e}")
        return set()


def search_in_pdf(pdf_path: str, config: dict) -> Set[str]:
    """Обработка PDF файлов"""
    if not HAS_PDF:
        return set()

    found = set()
    try:
        with fitz.open(pdf_path) as doc:
            for page in doc:
                # Текст со страницы
                text = page.get_text()
                found.update(search_in_text(text))

                # Обработка изображений (только если есть OCR)
                if HAS_OCR:
                    for img in page.get_images(full=True):
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        if base_image and "image" in base_image:
                            image_data = BytesIO(base_image["image"])
                            found.update(search_in_image(image_data, config))
    except Exception as e:
        logging.error(f"Ошибка обработки PDF {pdf_path}: {e}")
    return found


def search_in_docx(docx_path: str, config: dict) -> Set[str]:
    """Обработка DOCX файлов"""
    if not HAS_DOCX:
        return set()

    found = set()
    try:
        # Текст из документа
        text = docx2txt.process(docx_path)
        found.update(search_in_text(text))

        # Изображения из документа (только если есть OCR)
        if HAS_OCR:
            with tempfile.TemporaryDirectory() as temp_dir:
                docx2txt.process(docx_path, temp_dir)
                for img_file in os.listdir(temp_dir):
                    if img_file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')):
                        img_path = os.path.join(temp_dir, img_file)
                        found.update(search_in_image(img_path, config))
    except Exception as e:
        logging.error(f"Ошибка обработки DOCX {docx_path}: {e}")
    return found


def search_in_excel(excel_path: str) -> Set[str]:
    """Обработка Excel файлов"""
    if not HAS_EXCEL:
        return set()

    found = set()
    try:
        # Пропускаем временные файлы Excel
        if os.path.basename(excel_path).startswith('~$'):
            return set()

        if excel_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            found.update(search_in_text(cell))
        else:  # .xls
            df = pd.read_excel(excel_path, sheet_name=None)
            for sheet_name, sheet_data in df.items():
                for _, row in sheet_data.iterrows():
                    for value in row:
                        if isinstance(value, str):
                            found.update(search_in_text(str(value)))
    except Exception as e:
        logging.error(f"Ошибка обработки Excel {excel_path}: {e}")
    return found


def search_in_archive(archive_path: str, extensions: List[str]) -> Set[str]:
    """Обработка архивов"""
    found = set()
    try:
        if archive_path.endswith('.zip'):
            with zipfile.ZipFile(archive_path, 'r') as z:
                for file in z.namelist():
                    if any(fnmatch.fnmatch(file, ext) for ext in extensions):
                        with z.open(file) as f:
                            content = f.read().decode('utf-8', errors='ignore')
                            found.update(search_in_text(content))
        elif archive_path.endswith('.7z') and HAS_7Z:
            with py7zr.SevenZipFile(archive_path, mode='r') as z:
                archive_files = z.getnames()
                for file in archive_files:
                    if any(fnmatch.fnmatch(file, ext) for ext in extensions):
                        # Исправленный метод обработки 7z архивов
                        with tempfile.TemporaryDirectory() as temp_dir:
                            z.extract(path=temp_dir, targets=[file])
                            extracted_file = os.path.join(temp_dir, file)
                            if os.path.isfile(extracted_file):
                                try:
                                    with open(extracted_file, 'r', encoding='utf-8', errors='ignore') as f:
                                        content = f.read()
                                        found.update(search_in_text(content))
                                except:
                                    # Если не текстовый файл, пропускаем
                                    pass
        elif archive_path.endswith('.rar') and HAS_RAR:
            with rarfile.RarFile(archive_path, 'r') as z:
                for file in z.namelist():
                    if any(fnmatch.fnmatch(file, ext) for ext in extensions):
                        with z.open(file) as f:
                            content = f.read().decode('utf-8', errors='ignore')
                            found.update(search_in_text(content))
    except Exception as e:
        logging.error(f"Ошибка обработки архива {archive_path}: {e}")
    return found


def process_file(file_path: str, extensions: List[str], max_file_size: int, config: dict) -> Dict[str, Set[str]]:
    """Обработка отдельного файла"""
    found = set()
    try:
        # Проверяем размер файла
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        if file_size_mb > max_file_size:
            logging.warning(
                f"Пропуск файла {file_path} (размер {file_size_mb:.2f} МБ превышает лимит {max_file_size} МБ)")
            return {}

        ext = os.path.splitext(file_path)[1].lower()

        # Проверяем, соответствует ли файл заданным расширениям
        if not any(fnmatch.fnmatch(file_path, ext_pattern) for ext_pattern in extensions):
            return {}

        # Обработка в зависимости от типа файла
        if ext in ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff'):
            if HAS_OCR:
                found = search_in_image(file_path, config)
            else:
                logging.info(f"Пропуск изображения {file_path} (OCR недоступен)")
        elif ext == '.pdf' and HAS_PDF:
            found = search_in_pdf(file_path, config)
        elif ext == '.docx' and HAS_DOCX:
            found = search_in_docx(file_path, config)
        elif ext in ('.xls', '.xlsx') and HAS_EXCEL:
            found = search_in_excel(file_path)
        elif ext in ('.zip', '.7z', '.rar'):
            found = search_in_archive(file_path, extensions)
        else:
            # Обработка текстовых файлов
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    found = search_in_text(content)
            except UnicodeDecodeError:
                # Если UTF-8 не работает, пробуем другие кодировки
                encodings = ['cp1251', 'iso-8859-1', 'latin1']
                for encoding in encodings:
                    try:
                        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                            content = f.read()
                            found = search_in_text(content)
                            break
                    except UnicodeDecodeError:
                        continue

        return {file_path: found} if found else {}
    except Exception as e:
        logging.error(f"Ошибка обработки файла {file_path}: {e}")
        return {}


def search_files(root_dir: str, extensions: List[str], max_workers: int = 4, output_file: str = None,
                 max_file_size: int = 10, config: dict = None) -> Dict[str, Set[str]]:
    """Многопоточный поиск файлов"""
    results = {}

    # Собираем все файлы для обработки
    files_to_process = []
    for root, _, files in os.walk(root_dir):
        for file in files:
            file_path = os.path.join(root, file)
            if any(fnmatch.fnmatch(file, ext_pattern) for ext_pattern in extensions):
                files_to_process.append(file_path)

    logging.info(f"Найдено файлов для обработки: {len(files_to_process)}")

    # Открываем файл для записи результатов
    output_handle = None
    if output_file:
        output_handle = open(output_file, 'w', encoding='utf-8')
        output_handle.write("Результаты поиска:\n\n")
        output_handle.write(f"Время начала: {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n")

    # Обрабатываем файлы в несколько потоков с прогрессбаром
    with tqdm(total=len(files_to_process), desc="Обработка файлов", unit="файл") as pbar:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {
                executor.submit(process_file, file_path, extensions, max_file_size, config): file_path
                for file_path in files_to_process
            }

            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    if result:
                        results.update(result)
                        # Записываем результат в файл сразу
                        if output_handle:
                            for path, keywords_found in result.items():
                                output_handle.write(f"Файл: {path}\n")
                                output_handle.write(f"Найденные ключевые слова: {', '.join(keywords_found)}\n\n")
                                output_handle.flush()  # Принудительно записываем в файл
                except Exception as e:
                    logging.error(f"Ошибка при обработке файла {file_path}: {e}")
                finally:
                    # Обновляем прогрессбар
                    pbar.update(1)
                    pbar.set_postfix(file=os.path.basename(file_path)[:20])  # Показываем имя текущего файла

    # Закрываем файл результатов
    if output_handle:
        output_handle.write(f"\nВремя завершения: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        output_handle.close()

    return results


def create_default_config():
    """Создание файла конфигурации по умолчанию"""
    config_content = """[Settings]
# Расширения файлов для поиска (через запятую)
extensions = *.txt, *.pdf, *.docx, *.xlsx, *.jpg, *.png, *.zip, *.rar, *.7z

# Файл с ключевыми словами (каждое слово с новой строки)
keywords_file = keywords.txt

# Директория для поиска
directory = .

# Количество потоков для обработки
threads = 4

# Файл для сохранения результатов
output_file = search_results.txt

# Поиск по изображениям (требует установленного Tesseract OCR)
search_images = true

# Максимальный размер обрабатываемого файла (МБ)
max_file_size = 50

# Файл для логирования
log_file = search_log.txt

# Настройки Tesseract OCR
tesseract_languages = rus
tesseract_config = --oem 3 --psm 6
"""

    with open("config.txt", "w", encoding="utf-8") as f:
        f.write(config_content)

    logging.info("Создан файл конфигурации config.txt с настройками по умолчанию.")


def main():
    # Загружаем конфигурацию
    config = load_config()

    # Настраиваем логирование (очищает предыдущие логи)
    setup_logging(config.get('log_file', 'search_log.txt'))

    # Проверяем, существует ли файл конфигурации
    if not os.path.exists("config.txt"):
        logging.warning("Файл конфигурации config.txt не найден.")
        create_default_config()
        logging.info("Пожалуйста, настройте config.txt и запустите программу снова.")
        return

    extensions = config['extensions']
    keywords_file = config['keywords_file']
    directory = config['directory']
    threads = config['threads']
    output_file = config['output_file']
    search_images = config['search_images']
    max_file_size = config['max_file_size']

    # Проверяем существование файла с ключевыми словами
    if not os.path.exists(keywords_file):
        logging.error(f"Файл с ключевыми словами '{keywords_file}' не найден.")
        logging.info("Пожалуйста, создайте файл keywords.txt с ключевыми словами (каждое слово с новой строки).")
        return

    # Проверяем существование директории для поиска
    if not os.path.exists(directory):
        logging.error(f"Директория '{directory}' не существует.")
        return

    try:
        keywords = load_keywords(keywords_file)
    except ValueError as e:
        logging.error(e)
        return

    if not keywords:
        logging.error("Не найдено ключевых слов.")
        return

    logging.info(f"Загружена конфигурация из config.txt")
    logging.info(f"Директория для поиска: {directory}")
    logging.info(f"Используемые маски: {extensions}")
    logging.info(f"Файл с ключевыми словами: {keywords_file}")
    logging.info(f"Количество ключевых слов: {len(keywords)}")
    logging.info(f"Используется потоков: {threads}")
    logging.info(f"Файл для результатов: {output_file}")
    logging.info(f"Максимальный размер файла: {max_file_size} МБ")
    logging.info(f"Поиск по изображениям: {'включен' if search_images and HAS_OCR else 'отключен'}")

    # Выводим информацию о доступных функциях
    if not HAS_PDF:
        logging.warning("Обработка PDF недоступна")
    if not HAS_DOCX:
        logging.warning("Обработка DOCX недоступна")
    if not HAS_EXCEL:
        logging.warning("Обработка Excel недоступна")
    if not HAS_OCR and search_images:
        logging.warning("Поиск по изображениям недоступен (Tesseract не установлен)")
    if not HAS_7Z:
        logging.warning("Обработка 7z архивов недоступна")
    if not HAS_RAR:
        logging.warning("Обработка RAR архивов недоступна")

    # Выполняем поиск
    start_time = time.time()
    results = search_files(directory, extensions, threads, output_file, max_file_size, config)
    end_time = time.time()

    if results:
        logging.info(f"Найдено совпадений в {len(results)} файлах:")
        for file, found_keywords in results.items():
            logging.info(f"Файл: {file}")
            logging.info(f"Найденные ключевые слова: {', '.join(found_keywords)}")
        logging.info(f"Результаты также сохранены в файл {output_file}")
    else:
        logging.info("Ничего не найдено.")

    logging.info(f"Время выполнения: {end_time - start_time:.2f} секунд")


if __name__ == '__main__':
    main()